from django.shortcuts import render, redirect, get_object_or_404
from django.core.files.storage import FileSystemStorage
from openpyxl import load_workbook
from datetime import datetime

from django.http import HttpResponse

# Create your views here.

def index(request):
    # return HttpResponse("Hello World")
    return render(request, 'conversions/index.html')

def upload_xl(request):
    error_message = None
    if request.method == 'POST':
        if 'file' not in request.FILES:
            error_message = 'Please select an Excel (.xlsx) file to upload.'
            print(error_message)
            return render(request, 'conversions/upload_xl.html', {'error_message': error_message})
        else:
            uploaded_file = request.FILES['file'] 
            
            # Check if the uploaded file has a valid .xl extension
            if not uploaded_file.name.endswith('.xlsx'):
                error_message = 'Please upload a valid .xlsx file.'   
                return render(request, 'conversions/upload_xl.html', {'error_message': error_message})
                     
            # Save the uploaded file to a temporary location
            fs = FileSystemStorage()
            filename = fs.save(uploaded_file.name, uploaded_file)
            request.session['uploaded_file_path'] = fs.url(filename)
            return redirect('edit_file')
    
    return render(request, 'conversions/upload_xl.html')

def edit_file(request):
    if request.method == 'POST':        
        edited_data = []
        for key, value in request.POST.items():
            if key.startswith("cell_"):
                row_index = int(key.split('_')[1])
                edited_data.append({row_index : value})
                #print(int(key.split('_')[1]))
        #print(edited_data)
        
        # Initialize dictionary to store values for each rowIndex
        edited_table_dict = {}

        # Populate array_B_dict based on rowIndex
        for item in edited_data:
            for key, value in item.items():
                row_index = key  # Assuming the key represents the rowIndex
                if row_index not in edited_table_dict:
                    edited_table_dict[row_index] = []
            edited_table_dict[row_index].append(value)

        # Convert the dictionary into a list of lists
        edited_table = list(edited_table_dict.values())        
     
        request.session['edited_table'] = edited_table       
        #print(edited_data)
        print(edited_table)
        
        return redirect('convert_file')
    
    if 'uploaded_file_path' in request.session:
        uploaded_file_path = request.session.get('uploaded_file_path')
        wb = load_workbook(uploaded_file_path[1:])  # Remove the leading '/'
        sheet = wb.active    
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))
        #print(data)
        request.session.pop('uploaded_file_path', None)    
        return render(request, 'conversions/edit_file.html', {'data' : data})
    
    elif 'edited_list' in request.session:
        edited_list = request.session.get('edited_list')
        request.session.pop('edited_list', None)
        return render(request, 'conversions/edit_file.html', {'data' : edited_list})
    
def convert_file(request):
    edited_file = request.session.get('edited_table') 
    
    if not edited_file:
        return redirect('edit_file')
    
    # Transform data and create .feed content
    feed_content = []
    for row in edited_file[1:]:
        feed_entry = '|'.join(str(cell) for cell in row)  # Join data with '|' delimiter
        feed_content.append(feed_entry)
    
    
    
    if request.method == 'POST':
        # Write .feed content to a file
        vendor_name = "F1"
        country_name = "india"
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        print(timestamp)
        feed_file = f'INVENTORY_{vendor_name}_{country_name}_{timestamp}.feed'        
        with open(feed_file, 'w') as f:
            f.write('\n'.join(feed_content))

        # Create a response to download the file
        with open(feed_file, 'rb') as file:
            response = HttpResponse(file.read(), content_type='application/octet-stream')
            response['Content-Disposition'] = f'attachment; filename={feed_file}'

        print(f'Data converted and saved to {feed_file}')
        return response    
        
        
    return render(request, 'conversions/convert.html', {'data' : feed_content}) 